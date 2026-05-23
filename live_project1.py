import io
import math
import os
from dataclasses import dataclass
from typing import Dict, List, Optional

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill


try:
    import google.generativeai as genai
except ImportError:
    genai = None


st.set_page_config(
    page_title="Live Project 1 - AI Budget Planner",
    page_icon=":bar_chart:",
    layout="wide",
)


PALETTE = {
    "bg": "#F8FAFC",
    "panel": "#FFFFFF",
    "text": "#0F172A",
    "muted": "#475569",
    "line": "#D9E2EC",
    "accent": "#0EA5E9",
    "accent_soft": "#E0F2FE",
    "warning": "#F59E0B",
    "danger": "#EF4444",
    "success": "#16A34A",
}


DEFAULT_EXPENSES = [
    {"category": "Rent", "amount": 18000.0},
    {"category": "Subscriptions", "amount": 1200.0},
    {"category": "Self Care", "amount": 2500.0},
    {"category": "Miscellaneous", "amount": 3000.0},
]


@dataclass
class BudgetSnapshot:
    income: float
    total_expenses: float
    savings: float
    savings_rate: float
    needs: float
    wants: float
    investments: float


def inject_styles() -> None:
    st.markdown(
        f"""
        <style>
        .stApp {{
            background:
                radial-gradient(circle at 0% 0%, rgba(253, 224, 71, 0.18), transparent 22%),
                radial-gradient(circle at 100% 0%, rgba(14, 165, 233, 0.14), transparent 26%),
                linear-gradient(180deg, #FFFDF7 0%, #F8FBFF 52%, #F4F9FF 100%);
            color: {PALETTE["text"]};
        }}
        .block-container {{
            padding-top: 2rem;
            padding-bottom: 2rem;
            max-width: 1180px;
        }}
        .hero-card, .metric-card, .panel-card {{
            background: rgba(255, 255, 255, 0.9);
            border: 1px solid {PALETTE["line"]};
            border-radius: 18px;
            box-shadow: 0 20px 60px rgba(15, 23, 42, 0.08);
        }}
        .hero-card {{
            padding: 26px 28px;
            margin-bottom: 18px;
        }}
        .panel-card {{
            padding: 18px 18px 10px 18px;
        }}
        .metric-card {{
            padding: 16px 18px;
        }}
        .eyebrow {{
            display: inline-flex;
            align-items: center;
            gap: 8px;
            margin-bottom: 10px;
            color: {PALETTE["accent"]};
            font-size: 0.78rem;
            font-weight: 800;
            letter-spacing: 0.12em;
            text-transform: uppercase;
        }}
        .live-dot {{
            width: 10px;
            height: 10px;
            border-radius: 999px;
            background: #F87171;
            box-shadow: 0 0 0 6px rgba(248, 113, 113, 0.16);
        }}
        .hero-title {{
            margin: 0;
            color: {PALETTE["text"]};
            font-size: clamp(2rem, 4vw, 3.5rem);
            line-height: 1.02;
            font-weight: 800;
        }}
        .hero-copy {{
            margin-top: 12px;
            max-width: 780px;
            color: {PALETTE["muted"]};
            font-size: 1rem;
        }}
        .section-title {{
            margin: 0 0 10px 0;
            font-size: 1.15rem;
            font-weight: 700;
            color: {PALETTE["text"]};
        }}
        .small-note {{
            margin-top: 8px;
            color: {PALETTE["muted"]};
            font-size: 0.88rem;
        }}
        .metric-label {{
            color: {PALETTE["muted"]};
            font-size: 0.85rem;
            font-weight: 600;
        }}
        .metric-value {{
            color: {PALETTE["text"]};
            font-size: 1.45rem;
            font-weight: 800;
        }}
        .ai-box {{
            padding: 14px 16px;
            background: {PALETTE["accent_soft"]};
            border: 1px solid #BAE6FD;
            border-radius: 14px;
            color: {PALETTE["text"]};
            margin-bottom: 12px;
        }}
        div[data-testid="stDataFrame"] {{
            border: 1px solid {PALETTE["line"]};
            border-radius: 14px;
            overflow: hidden;
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )


def init_expenses() -> None:
    if "expense_rows" not in st.session_state:
        st.session_state.expense_rows = [row.copy() for row in DEFAULT_EXPENSES]


def add_expense_row() -> None:
    st.session_state.expense_rows.append({"category": "", "amount": 0.0})


def remove_expense_row(index: int) -> None:
    if len(st.session_state.expense_rows) > 1:
        st.session_state.expense_rows.pop(index)


def normalize_expenses(expenses: List[Dict[str, float]]) -> pd.DataFrame:
    cleaned: List[Dict[str, float]] = []
    for row in expenses:
        category = str(row.get("category", "")).strip() or "Uncategorized"
        try:
            amount = float(row.get("amount", 0) or 0)
        except (TypeError, ValueError):
            amount = 0.0
        cleaned.append({"Category": category, "Amount": max(amount, 0.0)})
    return pd.DataFrame(cleaned)


def classify_expense(category: str) -> str:
    normalized = category.lower()
    needs_keywords = {
        "rent",
        "groceries",
        "food",
        "utilities",
        "transport",
        "emi",
        "insurance",
        "medical",
        "fees",
        "phone",
        "internet",
    }
    wants_keywords = {
        "subscription",
        "entertainment",
        "shopping",
        "self care",
        "miscellaneous",
        "travel",
        "eating out",
        "fun",
    }
    if any(keyword in normalized for keyword in needs_keywords):
        return "Needs"
    if any(keyword in normalized for keyword in wants_keywords):
        return "Wants"
    return "Other"


def build_snapshot(income: float, expenses_df: pd.DataFrame) -> BudgetSnapshot:
    total_expenses = float(expenses_df["Amount"].sum())
    savings = income - total_expenses
    savings_rate = (savings / income * 100) if income > 0 else 0.0

    classified = expenses_df.copy()
    classified["Bucket"] = classified["Category"].apply(classify_expense)

    needs = float(classified.loc[classified["Bucket"] == "Needs", "Amount"].sum())
    wants = float(classified.loc[classified["Bucket"] == "Wants", "Amount"].sum())
    remaining = max(savings, 0.0)
    investments = max(remaining * 0.65, 0.0)

    return BudgetSnapshot(
        income=income,
        total_expenses=total_expenses,
        savings=savings,
        savings_rate=savings_rate,
        needs=needs,
        wants=wants,
        investments=investments,
    )


def build_summary_table(snapshot: BudgetSnapshot) -> pd.DataFrame:
    emergency_fund = max(snapshot.income * 0.1, 0.0)
    rd_target = max(snapshot.savings * 0.35, 0.0)
    sip_target = max(snapshot.savings * 0.45, 0.0)
    flexible_cash = max(snapshot.savings - rd_target - sip_target, 0.0)

    return pd.DataFrame(
        [
            {"Section": "Income", "Amount": snapshot.income},
            {"Section": "Total Expenses", "Amount": snapshot.total_expenses},
            {"Section": "Monthly Savings", "Amount": snapshot.savings},
            {"Section": "Emergency Buffer Goal", "Amount": emergency_fund},
            {"Section": "RD Allocation", "Amount": rd_target},
            {"Section": "SIP / Mutual Fund Bucket", "Amount": sip_target},
            {"Section": "Flexible Surplus", "Amount": flexible_cash},
        ]
    )


def build_rule_based_guidance(snapshot: BudgetSnapshot, expenses_df: pd.DataFrame) -> Dict[str, str]:
    top_row = expenses_df.sort_values("Amount", ascending=False).iloc[0]
    savings_band = "strong" if snapshot.savings_rate >= 25 else "moderate" if snapshot.savings_rate >= 10 else "tight"

    narrative = (
        f"Your biggest spending category is {top_row['Category']} at Rs. {top_row['Amount']:,.0f}. "
        f"Your current savings rate is {snapshot.savings_rate:.1f}%, which puts this month in a {savings_band} savings zone."
    )

    if snapshot.savings <= 0:
        action = (
            "Expenses are currently above income. The first move should be trimming flexible categories "
            "such as subscriptions, self-care, or miscellaneous spends before planning any investment bucket."
        )
    elif snapshot.savings_rate < 10:
        action = (
            "Try building a small emergency buffer first, then route the remaining surplus into disciplined saving options "
            "like an RD. Once your monthly cushion improves, you can gradually add a SIP or beginner-friendly mutual fund bucket."
        )
    else:
        action = (
            "You have room to split savings into an emergency buffer, a fixed saving bucket like an RD, "
            "and a long-term investing bucket such as SIPs or broad beginner-oriented mutual funds."
        )

    layout_brief = (
        "Excel layout brief: start with a clean summary banner, place the input table on the left, "
        "show a category breakdown chart beside it, then add savings guidance and next-step notes below."
    )

    return {
        "narrative": narrative,
        "action": action,
        "layout_brief": layout_brief,
    }


def get_gemini_guidance(snapshot: BudgetSnapshot, expenses_df: pd.DataFrame) -> Optional[Dict[str, str]]:
    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key or genai is None:
        return None

    prompt = f"""
    You are helping produce a budget planning report for a beginner-friendly personal finance demo.
    Stay general and safe. Do not name specific stocks, funds, or financial products.
    Focus on spending balance, savings habit, RD/SIP style buckets, and emergency fund logic.

    Income: {snapshot.income}
    Total expenses: {snapshot.total_expenses}
    Savings: {snapshot.savings}
    Savings rate: {snapshot.savings_rate:.2f}

    Expense table:
    {expenses_df.to_dict(orient="records")}

    Return three short sections with labels:
    NARRATIVE:
    ACTION:
    LAYOUT_BRIEF:
    """.strip()

    try:
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel("gemini-2.0-flash")
        response = model.generate_content(prompt)
        text = getattr(response, "text", "") or ""
        parsed = {"narrative": "", "action": "", "layout_brief": ""}
        current_key = None
        for raw_line in text.splitlines():
            line = raw_line.strip()
            upper = line.upper()
            if upper.startswith("NARRATIVE:"):
                current_key = "narrative"
                parsed[current_key] = line.split(":", 1)[1].strip()
            elif upper.startswith("ACTION:"):
                current_key = "action"
                parsed[current_key] = line.split(":", 1)[1].strip()
            elif upper.startswith("LAYOUT_BRIEF:"):
                current_key = "layout_brief"
                parsed[current_key] = line.split(":", 1)[1].strip()
            elif current_key and line:
                parsed[current_key] = f"{parsed[current_key]} {line}".strip()
        if all(parsed.values()):
            return parsed
    except Exception:
        return None
    return None


def create_excel_workbook(
    expenses_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    guidance: Dict[str, str],
) -> bytes:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Budget Summary"
    expenses_sheet = workbook.create_sheet("Expenses")
    notes_sheet = workbook.create_sheet("AI Notes")

    title_fill = PatternFill("solid", fgColor="DFF3FF")
    accent_fill = PatternFill("solid", fgColor="FFF4CC")
    header_font = Font(bold=True, color="0F172A")
    title_font = Font(bold=True, size=14, color="0F172A")

    summary_sheet["A1"] = "AI Budget Planner"
    summary_sheet["A1"].font = title_font
    summary_sheet["A2"] = "AI-assisted layout applied through Python workbook generation"
    summary_sheet["A4"] = "Section"
    summary_sheet["B4"] = "Amount"
    summary_sheet["A4"].fill = title_fill
    summary_sheet["B4"].fill = title_fill
    summary_sheet["A4"].font = header_font
    summary_sheet["B4"].font = header_font

    for idx, row in enumerate(summary_df.itertuples(index=False), start=5):
        summary_sheet.cell(row=idx, column=1, value=row.Section)
        summary_sheet.cell(row=idx, column=2, value=float(row.Amount))
        summary_sheet.cell(row=idx, column=2).number_format = '"Rs." #,##0.00'

    summary_sheet["D4"] = "AI Narrative"
    summary_sheet["D4"].fill = accent_fill
    summary_sheet["D4"].font = header_font
    summary_sheet["D5"] = guidance["narrative"]
    summary_sheet["D7"] = "Recommended Next Step"
    summary_sheet["D7"].fill = accent_fill
    summary_sheet["D7"].font = header_font
    summary_sheet["D8"] = guidance["action"]

    for col, width in {"A": 28, "B": 18, "D": 58}.items():
        summary_sheet.column_dimensions[col].width = width

    expenses_sheet["A1"] = "Expense Input Table"
    expenses_sheet["A1"].font = title_font
    expenses_sheet["A3"] = "Category"
    expenses_sheet["B3"] = "Amount"
    expenses_sheet["A3"].fill = title_fill
    expenses_sheet["B3"].fill = title_fill
    expenses_sheet["A3"].font = header_font
    expenses_sheet["B3"].font = header_font

    for idx, row in enumerate(expenses_df.itertuples(index=False), start=4):
        expenses_sheet.cell(row=idx, column=1, value=row.Category)
        expenses_sheet.cell(row=idx, column=2, value=float(row.Amount))
        expenses_sheet.cell(row=idx, column=2).number_format = '"Rs." #,##0.00'

    expenses_sheet.column_dimensions["A"].width = 30
    expenses_sheet.column_dimensions["B"].width = 16

    pie_chart = PieChart()
    labels = Reference(expenses_sheet, min_col=1, min_row=4, max_row=3 + len(expenses_df))
    data = Reference(expenses_sheet, min_col=2, min_row=3, max_row=3 + len(expenses_df))
    pie_chart.add_data(data, titles_from_data=True)
    pie_chart.set_categories(labels)
    pie_chart.title = "Expense Breakdown"
    pie_chart.height = 8
    pie_chart.width = 11
    expenses_sheet.add_chart(pie_chart, "D3")

    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.style = 10
    bar_chart.title = "Budget Summary"
    bar_chart.y_axis.title = "Amount"
    bar_chart.x_axis.title = "Section"
    bar_data = Reference(summary_sheet, min_col=2, min_row=4, max_row=7)
    bar_labels = Reference(summary_sheet, min_col=1, min_row=5, max_row=7)
    bar_chart.add_data(bar_data, titles_from_data=True)
    bar_chart.set_categories(bar_labels)
    bar_chart.height = 8
    bar_chart.width = 13
    summary_sheet.add_chart(bar_chart, "A14")

    notes_sheet["A1"] = "AI Export Layout Brief"
    notes_sheet["A1"].font = title_font
    notes_sheet["A3"] = guidance["layout_brief"]
    notes_sheet["A5"] = "Safety Note"
    notes_sheet["A5"].font = header_font
    notes_sheet["A6"] = (
        "This workbook keeps investment guidance general. It uses categories like RD, "
        "emergency fund, SIP, and beginner-friendly mutual fund buckets without naming specific products."
    )
    notes_sheet.column_dimensions["A"].width = 100

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def currency(value: float) -> str:
    prefix = "-" if value < 0 else ""
    return f"{prefix}Rs. {abs(value):,.0f}"


inject_styles()
init_expenses()

st.markdown(
    """
    <div class="hero-card">
      <div class="eyebrow"><span class="live-dot"></span>Live Project 1</div>
      <h1 class="hero-title">AI Budget Planning Assistant</h1>
      <p class="hero-copy">
        A Streamlit-based live demo for structured budgeting, safe savings guidance, category-level analytics,
        and AI-assisted Excel export planning.
      </p>
    </div>
    """,
    unsafe_allow_html=True,
)

left, right = st.columns([1.05, 0.95], gap="large")

with left:
    st.markdown('<div class="panel-card">', unsafe_allow_html=True)
    st.markdown('<h3 class="section-title">Income and Monthly Spends</h3>', unsafe_allow_html=True)
    monthly_income = st.number_input("Monthly earnings", min_value=0.0, step=1000.0, value=50000.0)

    st.markdown("#### Expense rows")
    for index, row in enumerate(st.session_state.expense_rows):
        col1, col2, col3 = st.columns([2.2, 1.2, 0.5])
        category = col1.text_input(
            f"Category {index + 1}",
            value=row["category"],
            key=f"category_{index}",
            label_visibility="collapsed",
            placeholder="Expense category",
        )
        amount = col2.number_input(
            f"Amount {index + 1}",
            min_value=0.0,
            step=500.0,
            value=float(row["amount"]),
            key=f"amount_{index}",
            label_visibility="collapsed",
        )
        remove_clicked = col3.button("X", key=f"remove_{index}", use_container_width=True)
        st.session_state.expense_rows[index]["category"] = category
        st.session_state.expense_rows[index]["amount"] = amount
        if remove_clicked:
            remove_expense_row(index)
            st.rerun()

    add_col, run_col = st.columns([1, 1])
    add_col.button("+ Add expense", on_click=add_expense_row, use_container_width=True)
    process_clicked = run_col.button("Process budget", type="primary", use_container_width=True)
    st.markdown(
        '<p class="small-note">Version 1 works without a database. Everything is calculated per request and can be exported directly.</p>',
        unsafe_allow_html=True,
    )
    st.markdown("</div>", unsafe_allow_html=True)

expenses_df = normalize_expenses(st.session_state.expense_rows)
snapshot = build_snapshot(monthly_income, expenses_df)
summary_df = build_summary_table(snapshot)
guidance = get_gemini_guidance(snapshot, expenses_df) or build_rule_based_guidance(snapshot, expenses_df)
excel_bytes = create_excel_workbook(expenses_df, summary_df, guidance)

with right:
    metric_cols = st.columns(3)
    metric_cards = [
        ("Savings", currency(snapshot.savings)),
        ("Savings Rate", f"{snapshot.savings_rate:.1f}%"),
        ("Top Bucket", expenses_df.sort_values("Amount", ascending=False).iloc[0]["Category"]),
    ]
    for col, (label, value) in zip(metric_cols, metric_cards):
        with col:
            st.markdown(
                f"""
                <div class="metric-card">
                  <div class="metric-label">{label}</div>
                  <div class="metric-value">{value}</div>
                </div>
                """,
                unsafe_allow_html=True,
            )

    st.markdown("")
    st.markdown('<div class="panel-card">', unsafe_allow_html=True)
    st.markdown('<h3 class="section-title">AI Summary</h3>', unsafe_allow_html=True)
    st.markdown(f'<div class="ai-box"><strong>Spending read:</strong> {guidance["narrative"]}</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="ai-box"><strong>Planning direction:</strong> {guidance["action"]}</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="ai-box"><strong>Excel layout brief:</strong> {guidance["layout_brief"]}</div>', unsafe_allow_html=True)
    st.download_button(
        "Download Excel budget pack",
        data=excel_bytes,
        file_name="ai_budget_planner.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    st.markdown("</div>", unsafe_allow_html=True)

if process_clicked or True:
    st.markdown("")
    table_col, chart_col = st.columns([1, 1], gap="large")

    with table_col:
        st.markdown('<h3 class="section-title">Budget Summary Table</h3>', unsafe_allow_html=True)
        display_summary = summary_df.copy()
        display_summary["Amount"] = display_summary["Amount"].map(currency)
        st.dataframe(display_summary, use_container_width=True, hide_index=True)

        st.markdown('<h3 class="section-title">Expense Table</h3>', unsafe_allow_html=True)
        display_expenses = expenses_df.copy()
        display_expenses["Amount"] = display_expenses["Amount"].map(currency)
        st.dataframe(display_expenses, use_container_width=True, hide_index=True)

    with chart_col:
        st.markdown('<h3 class="section-title">Expense Category Breakdown</h3>', unsafe_allow_html=True)
        st.bar_chart(expenses_df.set_index("Category"))

        st.markdown('<h3 class="section-title">Spend vs Save</h3>', unsafe_allow_html=True)
        spend_save_df = pd.DataFrame(
            {
                "Type": ["Expenses", "Savings"],
                "Amount": [snapshot.total_expenses, max(snapshot.savings, 0.0)],
            }
        ).set_index("Type")
        st.bar_chart(spend_save_df)

        if snapshot.savings > 0:
            allocation_df = pd.DataFrame(
                {
                    "Bucket": ["RD", "SIP / Mutual Fund Bucket", "Flexible Cash"],
                    "Amount": [
                        max(snapshot.savings * 0.35, 0.0),
                        max(snapshot.savings * 0.45, 0.0),
                        max(snapshot.savings * 0.20, 0.0),
                    ],
                }
            ).set_index("Bucket")
            st.markdown('<h3 class="section-title">Suggested Savings Split</h3>', unsafe_allow_html=True)
            st.bar_chart(allocation_df)

st.caption(
    "AI is used here for explanation and export-structure guidance. Calculations and the final Excel workbook are generated deterministically in Python."
)
